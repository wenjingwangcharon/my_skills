#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把帖子数据生成 notes.xlsx：每行一个帖子，元数据列在前，图片以“嵌入单元格”
的方式放在最后几列（image_1..image_N）。图片转 JPEG 内嵌，xlsx 自包含。

用法：
  python build_xlsx.py --data output/xhs_data.json --xlsx output/notes.xlsx
"""
import argparse
import io
import json
import os
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

BEIJING = timezone(timedelta(hours=8))


def fmt_time(v):
    if v is None:
        return ""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if v > 1e11:
        v = v / 1000.0
    elif v <= 0:
        return str(v)
    try:
        dt = datetime.fromtimestamp(v, tz=timezone.utc).astimezone(BEIJING)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(v)


THUMB = 110
IMG_W = 14
IMG_H = 82
HEADERS = ["note_id", "type", "title", "desc", "likes", "collected",
           "comment_count", "share_count", "time", "url", "image_count"]


def load_thumb(path):
    with PILImage.open(path) as im:
        im = im.convert("RGB")
        im.thumbnail((THUMB, THUMB))
        buf = io.BytesIO()
        im.save(buf, "JPEG", quality=85)
        buf.seek(0)
        return buf, im.width, im.height


def run(data_file, xlsx_path):
    d = json.load(open(data_file, encoding="utf-8"))
    notes = d.get("notes", [])
    max_imgs = max((n.get("image_count", 0) for n in notes), default=0)
    img_cols = [f"image_{i+1}" for i in range(max_imgs)]

    wb = Workbook()
    ws = wb.active
    ws.title = "notes"

    header_fill = PatternFill("solid", fgColor="FFE5B4")
    header_font = Font(bold=True)
    headers = HEADERS + img_cols
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    img_start = len(HEADERS) + 1  # 图片列起始列号
    for r, n in enumerate(notes, start=2):
        ii = n.get("interact_info") or {}
        meta = [
            n.get("note_id"),
            n.get("type"),
            (n.get("title") or "").replace("\n", " "),
            (n.get("desc") or "").replace("\n", " "),
            ii.get("liked_count"),
            ii.get("collected_count"),
            ii.get("comment_count"),
            ii.get("share_count"),
            n.get("time_text") or fmt_time(n.get("time")),
            f"https://www.xiaohongshu.com/explore/{n.get('note_id')}",
            n.get("image_count", 0),
        ]
        for c, v in enumerate(meta, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3, 4):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

        locals_ = [p for p in (n.get("image_local") or []) if p]
        for i, rel in enumerate(locals_):
            fpath = os.path.join(os.path.dirname(os.path.abspath(data_file)), rel)
            if not os.path.exists(fpath):
                continue
            try:
                buf, w, h = load_thumb(fpath)
                img = XLImage(buf)
                scale = min(THUMB / w, THUMB / h, 1.0)
                img.width = int(w * scale)
                img.height = int(h * scale)
                col = get_column_letter(img_start + i)
                ws.add_image(img, f"{col}{r}")
            except Exception as e:
                print(f"    [warn] 图片嵌入失败 {rel}: {e}")

        ws.row_dimensions[r].height = max(IMG_H, THUMB * 0.75)

    widths = {"A": 26, "B": 6, "C": 28, "D": 50, "E": 8, "F": 9,
              "G": 12, "H": 10, "I": 12, "J": 40, "K": 11}
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt
    for i in range(max_imgs):
        ws.column_dimensions[get_column_letter(img_start + i)].width = IMG_W

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
    wb.save(xlsx_path)
    print(f"已生成: {xlsx_path}")
    print(f"行数(含表头): {len(notes)+1} | 图片列: {max_imgs} | 嵌入图片总数: {sum(n.get('image_count',0) for n in notes)}")


def main():
    ap = argparse.ArgumentParser(description="生成嵌入图片的 notes.xlsx")
    ap.add_argument("--data", default=os.path.join(os.getcwd(), "output", "xhs_data.json"))
    ap.add_argument("--xlsx", default=None, help="输出 xlsx 路径，默认 <data目录>/notes.xlsx")
    args = ap.parse_args()
    if args.xlsx is None:
        args.xlsx = os.path.join(os.path.dirname(os.path.abspath(args.data)), "notes.xlsx")
    run(args.data, args.xlsx)


if __name__ == "__main__":
    main()
