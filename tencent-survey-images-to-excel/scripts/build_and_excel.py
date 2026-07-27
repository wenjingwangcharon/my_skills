#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成精简字段 + 嵌入图片缩略图的 Excel"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

records = json.load(open('/tmp/ima_imgs/final_records.json', encoding='utf-8'))
records.sort(key=lambda r: r['id'])

wb = Workbook()
ws = wb.active
ws.title = "投稿数据"

# 列定义
headers = ["编号", "结束答题时间", "ima UID", "发布平台", "发布内容链接",
           "联系方式", "省份", "城市", "个人主页截图", "发布内容截图"]
COL = {h: i+1 for i, h in enumerate(headers)}

# 样式
header_fill = PatternFill("solid", fgColor="2C3E50")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(vertical="center", horizontal="center")

# 写表头
for h, c in COL.items():
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border
ws.row_dimensions[1].height = 24

# 列宽
widths = {"编号":6, "结束答题时间":17, "ima UID":24, "发布平台":14,
          "发布内容链接":42, "联系方式":22, "省份":10, "城市":10,
          "个人主页截图":26, "发布内容截图":26}
for h, w in widths.items():
    ws.column_dimensions[get_column_letter(COL[h])].width = w

THUMB_W = 160  # 缩略图宽度(px)
PAD = 6

def add_thumb(path, col_letter, top_px, max_w=THUMB_W):
    """在指定列插入缩略图，返回该图占用高度(px)"""
    try:
        with PILImage.open(path) as im:
            w, h = im.size
    except Exception:
        return 0
    ratio = max_w / w
    disp_w = max_w
    disp_h = int(h * ratio)
    # 限制单图最大高度，避免超长截图
    MAX_H = 260
    if disp_h > MAX_H:
        disp_h = MAX_H
        disp_w = int(w * (MAX_H / h))
    img = XLImage(path)
    img.width = disp_w
    img.height = disp_h
    img.anchor = f"{col_letter}{cur_row}"
    # 用 offset 让多图纵向堆叠
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    col_idx = COL_LETTER_IDX[col_letter] - 1
    marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(PAD), row=cur_row-1, rowOff=pixels_to_EMU(top_px+PAD))
    size = XDRPositiveSize2D(pixels_to_EMU(disp_w), pixels_to_EMU(disp_h))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)
    return disp_h + PAD

COL_LETTER_IDX = {get_column_letter(i): i for i in range(1, 30)}

cur_row = 2
for rec in records:
    ws.cell(row=cur_row, column=COL["编号"], value=rec['id'])
    ws.cell(row=cur_row, column=COL["结束答题时间"], value=rec['ended_at'])
    ws.cell(row=cur_row, column=COL["ima UID"], value=rec['uid'])
    ws.cell(row=cur_row, column=COL["发布平台"], value=rec['platform'])
    ws.cell(row=cur_row, column=COL["发布内容链接"], value=rec['link'])
    ws.cell(row=cur_row, column=COL["联系方式"], value=rec['contact'])
    ws.cell(row=cur_row, column=COL["省份"], value=rec['province'])
    ws.cell(row=cur_row, column=COL["城市"], value=rec['city'])

    # 插入主页截图
    home_letter = get_column_letter(COL["个人主页截图"])
    content_letter = get_column_letter(COL["发布内容截图"])
    home_h = PAD
    for f in rec['home_files']:
        if f['local']:
            home_h += add_thumb(f['local'], home_letter, home_h)
    content_h = PAD
    for f in rec['content_files']:
        if f['local']:
            content_h += add_thumb(f['local'], content_letter, content_h)

    row_h_px = max(home_h, content_h, 40)
    ws.row_dimensions[cur_row].height = row_h_px * 0.75  # px -> points

    # 单元格样式
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=cur_row, column=c)
        cell.border = border
        if c in (COL["编号"], COL["发布平台"], COL["省份"], COL["城市"], COL["结束答题时间"]):
            cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
        else:
            cell.alignment = wrap
    cur_row += 1

ws.freeze_panes = "A2"

out = "/Users/lalalacharon/WorkBuddy/2026-06-26-11-33-54/运营活动-投稿数据(含图片).xlsx"
wb.save(out)
print("已保存:", out)
print("总记录:", len(records))
