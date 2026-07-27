#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""阶段二：把内容评分追加到「含图片的成品表」后面（同一张表，方便对照图片核查）。

用法：
  1. 阶段一已生成含图片的成品表（如 运营活动-投稿数据(含图片).xlsx）。
  2. 逐条访问内容链接打分后，把结果整理成 scores.json，结构见下方 SCORES_EXAMPLE。
  3. 改 SRC / SCORES_JSON 路径后运行：
     ~/.workbuddy/binaries/python/envs/default/bin/python append_scores.py

关键点：
  - 用 load_workbook 打开已含图片的成品表再追加列，绝不新建工作簿，否则丢图片。
  - 按 A 列"编号"逐行匹配评分。
  - 链接不可访问的记录：评分留空、状态写原因、整行评分区标浅红底色。
  - 综合分 = 三个子项均值（保留2位），任一子项为 None 则综合分为 None。
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===== 改这里 =====
SRC = "/Users/lalalacharon/WorkBuddy/2026-06-26-11-33-54/运营活动-投稿数据(含图片).xlsx"
SCORES_JSON = "/tmp/ima_imgs/scores.json"   # 评分数据（见下方结构）

# scores.json 结构示例（key 为编号字符串/数字均可）：
SCORES_EXAMPLE = {
    "3": {
        "status": "可访问",
        "wd":  [5, 5, 5],      # 内容质量三项：完整度/清晰度/实操性
        "rep": [4, 5, 4],      # 案例代表性三项：场景创新性/示范价值/共鸣度
        "note": "评分说明文字…"
    },
    "5": {
        "status": "不可访问-小红书短链失效/风控",
        "wd":  [None, None, None],
        "rep": [None, None, None],
        "note": "短链404，浏览器风控，无法查看。"
    },
}

# 追加的评分列（紧跟在图片列之后）
SCORE_HEADERS = ["状态", "完整度", "清晰度", "实操性", "内容质量综合分",
                 "场景创新性", "示范价值", "共鸣度", "案例代表性综合分", "评分说明"]


def avg3(vals):
    if any(v is None for v in vals):
        return None
    return round(sum(vals) / 3, 2)


def main():
    scores = json.load(open(SCORES_JSON, encoding="utf-8"))
    # 统一 key 为 str
    scores = {str(k): v for k, v in scores.items()}

    wb = load_workbook(SRC)
    ws = wb.active

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(color="FFFFFF", bold=True)
    na_fill = PatternFill("solid", fgColor="FFE0E0")

    start_col = ws.max_column + 1

    # 表头
    for i, h in enumerate(SCORE_HEADERS):
        c = ws.cell(row=1, column=start_col + i, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    # 数据行（按 A 列编号匹配）
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(row=r, column=1).value)
        s = scores.get(rid)
        if not s:
            continue
        wd, rep = s["wd"], s["rep"]
        row_vals = [s["status"], wd[0], wd[1], wd[2], avg3(wd),
                    rep[0], rep[1], rep[2], avg3(rep), s.get("note", "")]
        inaccessible = str(s["status"]).startswith("不可访问")
        for i, v in enumerate(row_vals):
            c = ws.cell(row=r, column=start_col + i, value=v)
            c.border = thin
            if i == 9:  # 说明列左对齐
                c.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if inaccessible:
                c.fill = na_fill

    # 列宽
    widths = {0: 24, 1: 8, 2: 8, 3: 8, 4: 13, 5: 11, 6: 9, 7: 8, 8: 14, 9: 55}
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(start_col + i)].width = w

    wb.save(SRC)
    print("已追加评分列到:", SRC)
    print("总列数:", ws.max_column, "| 图片数(应不变):", len(ws._images))


if __name__ == "__main__":
    main()
