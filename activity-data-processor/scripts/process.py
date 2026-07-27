import sys
import argparse
import json
import math
import subprocess
import pandas as pd

from openpyxl.formatting.rule import DataBar
from openpyxl.descriptors import Bool

if not hasattr(DataBar, 'gradient'):
    DataBar.gradient = Bool(allow_none=True)
    _orig_dbar_init = DataBar.__init__
    def _patched_dbar_init(self, minLength=None, maxLength=None, showValue=None,
                           cfvo=None, color=None, gradient=None):
        _orig_dbar_init(self, minLength=minLength, maxLength=maxLength,
                        showValue=showValue, cfvo=cfvo, color=color)
        self.gradient = gradient
    DataBar.__init__ = _patched_dbar_init

WECOM_DOCID = 'dcGEuIWvPh-j9gHyhf7TrR6UcgvtlNNXOdnB6VCccntCkipuiAo0I6CBF3m0BPxcNNt4vgTMA4zHsz5cQuSE-Duw'
WECOM_SHEET1_ID = 'BB08J2'
WECOM_SHEET2_ID = '7SIXPe'

ACTIVITY_ORDER = [
    '邀好友送算力', '知识号算力激励-领券', '知识号算力激励-兑换',
    '每日登录福利', '新用户首月福利', '新人福利',
    'copilot算力补贴', 'copilot用户奖励', '下载双端领算力'
]

SHEET1_HEADERS = [
    '活动', '本周下发算力', '下发占比', '环比变化', '本周下发人数', '本周人均下发',
    '本周会话消耗算力', '消耗占比', '环比变化', '本周消耗人数', '本周人均消耗',
    '本周过期消耗算力', '过期占比', '环比变化', '本周过期人数', '本周人均过期'
]

SHEET2_HEADERS = [
    '活动', '累计下发算力', '下发占比', '累计下发人数',
    '累计消耗算力', '消耗率', '累计会话消耗人数',
    '累计过期消耗算力', '过期率', '累计过期消耗人数',
    '余额', '余额率', '累计付费人数', '付费率'
]

S1_NCOLS = len(SHEET1_HEADERS)
S2_NCOLS = len(SHEET2_HEADERS)

CENTER = {'horizontal_alignment': 'HORIZONTAL_CENTER', 'vertical_alignment': 'VERTICAL_CENTER'}
BOLD_CENTER = {'horizontal_alignment': 'HORIZONTAL_CENTER', 'vertical_alignment': 'VERTICAL_CENTER', 'text_format': {'bold': True}}


def fmt_wan(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, str):
        if '万' in v:
            return v
        try:
            v = float(v.replace(',', ''))
        except ValueError:
            return v
    try:
        import numpy as np
        if isinstance(v, (np.integer, np.floating)):
            v = float(v)
    except ImportError:
        pass
    if isinstance(v, (int, float)):
        if v == 0:
            return '0'
        if abs(v) >= 10000:
            return f'{v/10000:.2f}万'
        return str(int(v)) if v == int(v) else f'{v:.0f}'
    return str(v)


def fmt_pct(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, str):
        try:
            v = float(v.replace(',', '').replace('%', ''))
        except ValueError:
            return v
    if isinstance(v, (int, float)):
        return f'{v*100:.2f}%'
    return str(v)


def fmt_int(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, str):
        try:
            v = float(v.replace(',', ''))
        except ValueError:
            return v
    try:
        import numpy as np
        if isinstance(v, np.integer):
            return str(int(v))
    except ImportError:
        pass
    if isinstance(v, (int, float)):
        return str(int(v))
    return str(v)


def safe_div(a, b):
    if b is None or b == 0:
        return None
    return a / b


def read_source(path):
    df = pd.read_excel(path, header=0)
    df = df[df['活动'] != '购买算力']
    total_row = df[df['sort_no'] == 999]
    detail_rows = df[df['sort_no'] != 999].set_index('活动')
    return detail_rows, total_row


def build_sheet1_block(detail, total):
    rows = []
    raw_mom = []

    total_issued = detail['本周下发算力'].sum()
    total_consumed = detail['本周会话消耗算力'].sum()
    total_expired = detail['本周过期消耗算力'].sum()
    total_lw_issued = detail['上周下发算力'].sum()
    total_lw_consumed = detail['上周消耗算力'].sum()
    total_lw_expired = detail['上周过期算力'].sum()

    total_issued_users = int(detail['本周下发人数'].fillna(0).sum())
    total_consumed_users = int(detail['本周消耗人数'].fillna(0).sum())
    total_expired_users = int(detail['本周过期人数'].fillna(0).sum())

    for act in ACTIVITY_ORDER:
        r = detail.loc[act]
        issued = r['本周下发算力']
        consumed = r['本周会话消耗算力']
        expired = r['本周过期消耗算力']
        lw_issued = r['上周下发算力']
        lw_consumed = r['上周消耗算力']
        lw_expired = r['上周过期算力']

        row = [
            act,
            fmt_wan(issued),
            fmt_pct(safe_div(issued, total_issued)),
            fmt_pct(safe_div(issued - lw_issued, lw_issued)),
            fmt_int(r['本周下发人数']) if pd.notna(r['本周下发人数']) else '0',
            fmt_int(round(issued / r['本周下发人数'])) if r['本周下发人数'] and r['本周下发人数'] > 0 else None,
            fmt_wan(consumed),
            fmt_pct(safe_div(consumed, total_consumed)),
            fmt_pct(safe_div(consumed - lw_consumed, lw_consumed)),
            fmt_int(r['本周消耗人数']) if pd.notna(r['本周消耗人数']) else '0',
            fmt_int(round(consumed / r['本周消耗人数'])) if r['本周消耗人数'] and r['本周消耗人数'] > 0 else None,
            fmt_wan(expired),
            fmt_pct(safe_div(expired, total_expired)),
            fmt_pct(safe_div(expired - lw_expired, lw_expired)),
            fmt_int(r['本周过期人数']) if pd.notna(r['本周过期人数']) else '0',
            fmt_int(round(expired / r['本周过期人数'])) if r['本周过期人数'] and r['本周过期人数'] > 0 else None,
        ]
        rows.append(row)
        raw_mom.append((
            safe_div(issued - lw_issued, lw_issued),
            safe_div(consumed - lw_consumed, lw_consumed),
            safe_div(expired - lw_expired, lw_expired)
        ))

    total_row_data = [
        'TOTAL',
        fmt_wan(total_issued),
        fmt_pct(1),
        fmt_pct(safe_div(total_issued - total_lw_issued, total_lw_issued)),
        fmt_int(total_issued_users),
        fmt_int(round(total_issued / total_issued_users)) if total_issued_users > 0 else None,
        fmt_wan(total_consumed),
        fmt_pct(1),
        fmt_pct(safe_div(total_consumed - total_lw_consumed, total_lw_consumed)),
        fmt_int(total_consumed_users),
        fmt_int(round(total_consumed / total_consumed_users)) if total_consumed_users > 0 else None,
        fmt_wan(total_expired),
        fmt_pct(1),
        fmt_pct(safe_div(total_expired - total_lw_expired, total_lw_expired)),
        fmt_int(total_expired_users),
        fmt_int(round(total_expired / total_expired_users)) if total_expired_users > 0 else None,
    ]
    rows.append(total_row_data)
    raw_mom.append((
        safe_div(total_issued - total_lw_issued, total_lw_issued),
        safe_div(total_consumed - total_lw_consumed, total_lw_consumed),
        safe_div(total_expired - total_lw_expired, total_lw_expired)
    ))
    raw_totals = {
        'issued': total_issued, 'consumed': total_consumed, 'expired': total_expired,
        'detail_issued': [detail.loc[act]['本周下发算力'] for act in ACTIVITY_ORDER],
        'detail_consumed': [detail.loc[act]['本周会话消耗算力'] for act in ACTIVITY_ORDER],
        'detail_expired': [detail.loc[act]['本周过期消耗算力'] for act in ACTIVITY_ORDER],
    }
    return rows, raw_mom, raw_totals


def build_sheet2_block(detail, total):
    rows = []
    total_c_issued = detail['累计下发算力'].sum()
    total_c_consumed = detail['累计会话消耗算力'].sum()
    total_c_expired = detail['累计过期消耗算力'].sum()
    total_c_issued_users = int(detail['累计下发人数'].fillna(0).sum())
    total_c_consumed_users = int(detail['累计会话消耗人数'].fillna(0).sum())
    total_c_expired_users = int(detail['累计过期人数'].fillna(0).sum())
    total_c_paid = int(detail['累计付费用户'].fillna(0).sum())

    for act in ACTIVITY_ORDER:
        r = detail.loc[act]
        c_issued = r['累计下发算力']
        c_consumed = r['累计会话消耗算力']
        c_expired = r['累计过期消耗算力']
        c_issued_users = r['累计下发人数']
        c_consumed_users = r['累计会话消耗人数']
        c_expired_users = r['累计过期人数']
        c_paid = r['累计付费用户']

        balance = c_issued - c_consumed - c_expired if pd.notna(c_issued) else 0
        row = [
            act,
            fmt_wan(c_issued),
            fmt_pct(safe_div(c_issued, total_c_issued)),
            fmt_int(c_issued_users) if pd.notna(c_issued_users) else '0',
            fmt_wan(c_consumed),
            fmt_pct(safe_div(c_consumed, c_issued)),
            fmt_int(c_consumed_users) if pd.notna(c_consumed_users) else '0',
            fmt_wan(c_expired),
            fmt_pct(safe_div(c_expired, c_issued)),
            fmt_int(c_expired_users) if pd.notna(c_expired_users) else '0',
            fmt_wan(balance),
            fmt_pct(safe_div(balance, c_issued)),
            fmt_int(c_paid) if pd.notna(c_paid) else '0',
            fmt_pct(safe_div(c_paid, c_issued_users)),
        ]
        rows.append(row)

    total_balance = total_c_issued - total_c_consumed - total_c_expired
    total_row_data = [
        'TOTAL',
        fmt_wan(total_c_issued),
        fmt_pct(1),
        fmt_int(total_c_issued_users),
        fmt_wan(total_c_consumed),
        fmt_pct(safe_div(total_c_consumed, total_c_issued)),
        fmt_int(total_c_consumed_users),
        fmt_wan(total_c_expired),
        fmt_pct(safe_div(total_c_expired, total_c_issued)),
        fmt_int(total_c_expired_users),
        fmt_wan(total_balance),
        fmt_pct(safe_div(total_balance, total_c_issued)),
        fmt_int(total_c_paid),
        fmt_pct(safe_div(total_c_paid, total_c_issued_users)),
    ]
    rows.append(total_row_data)
    return rows


def fmt_text(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return str(v)


def cell_to_text(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return str(v)


def _inject_negative_color(xlsx_path, neg_hex='E15759'):
    import re
    import shutil
    import zipfile
    tmp = xlsx_path + '.tmp'
    color = f'00{neg_hex}'
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                content = zin.read(item)
                if item.startswith('xl/worksheets/sheet') and item.endswith('.xml'):
                    text = content.decode('utf-8')
                    if 'dataBar' in text:
                        def _inject(m):
                            inner = m.group(0)
                            if 'negativeColor' in inner:
                                return inner
                            return re.sub(
                                r'(<color rgb="[0-9A-Fa-f]+"/>)',
                                rf'\1<negativeColor rgb="{color}"/>',
                                inner, count=1)
                        text = re.sub(
                            r'<dataBar[^>]*>.*?</dataBar>',
                            _inject, text, flags=re.DOTALL)
                        content = text.encode('utf-8')
                zout.writestr(item, content)
    shutil.move(tmp, xlsx_path)


def local_excel_to_wecom_rows(path, sheet_name, ncols, bold_rows, col_formatters):
    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    rows = []
    for ri in range(df.shape[0]):
        vals = []
        fmt = BOLD_CENTER if ri in bold_rows else CENTER
        for ci in range(min(df.shape[1], ncols)):
            raw = df.iloc[ri, ci]
            if ci < len(col_formatters):
                txt = col_formatters[ci](raw)
            else:
                txt = cell_to_text(raw)
            if txt is None:
                vals.append({'cell_format': fmt})
            else:
                vals.append({'cell_value': {'text': txt}, 'cell_format': fmt})
        for _ in range(max(0, ncols - df.shape[1])):
            vals.append({'cell_format': fmt})
        rows.append({'values': vals})
    return rows


def new_block_wecom_rows(date_str, headers, data_rows, ncols):
    rows = []

    vals = [{'cell_value': {'text': '本周数据'}, 'cell_format': BOLD_CENTER},
            {'cell_value': {'text': date_str}, 'cell_format': BOLD_CENTER}]
    for _ in range(ncols - 2):
        vals.append({'cell_format': BOLD_CENTER})
    rows.append({'values': vals})

    vals = []
    for h in headers:
        vals.append({'cell_value': {'text': h}, 'cell_format': BOLD_CENTER})
    rows.append({'values': vals})

    for ri, rdata in enumerate(data_rows):
        vals = []
        is_total = (ri == len(data_rows) - 1)
        fmt = BOLD_CENTER if is_total else CENTER
        for v in rdata:
            if v is None:
                vals.append({'cell_format': fmt})
            else:
                vals.append({'cell_value': {'text': str(v)}, 'cell_format': fmt})
        rows.append({'values': vals})

    for _ in range(2):
        vals = [{'cell_format': CENTER} for _ in range(ncols)]
        rows.append({'values': vals})

    return rows


def write_wecom(sheet_id, rows):
    payload = json.dumps({
        'docid': WECOM_DOCID,
        'sheet_id': sheet_id,
        'grid_data': {'start_row': 0, 'start_column': 0, 'rows': rows}
    }, ensure_ascii=False)
    result = subprocess.run(['wecom-cli', 'doc', 'sheet_update_range_data', payload],
                            capture_output=True, text=True, timeout=60)
    resp = json.loads(result.stdout)
    data = json.loads(resp['result']['content'][0]['text'])
    if data['errcode'] != 0:
        raise RuntimeError(f'WeCom write failed: {data}')


def process_wecom(source_path, date_str, local_output_path):
    detail, total = read_source(source_path)
    s1_new, _, _ = build_sheet1_block(detail, total)
    s2_new = build_sheet2_block(detail, total)

    S1_FORMATTERS = [fmt_text] + [fmt_wan, fmt_pct, fmt_pct, fmt_int, fmt_int, fmt_wan, fmt_pct, fmt_pct, fmt_int, fmt_int, fmt_wan, fmt_pct, fmt_pct, fmt_int, fmt_int]
    S2_FORMATTERS = [fmt_text] + [fmt_wan, fmt_pct, fmt_int, fmt_wan, fmt_pct, fmt_int, fmt_wan, fmt_pct, fmt_int, fmt_wan, fmt_pct, fmt_int, fmt_pct]

    s1_existing = local_excel_to_wecom_rows(local_output_path, 'Sheet1', S1_NCOLS, set(), S1_FORMATTERS)
    s2_existing = local_excel_to_wecom_rows(local_output_path, 'Sheet2', S2_NCOLS, set(), S2_FORMATTERS)

    s1_new_rows = new_block_wecom_rows(date_str, SHEET1_HEADERS, s1_new, S1_NCOLS)
    s2_new_rows = new_block_wecom_rows(date_str, SHEET2_HEADERS, s2_new, S2_NCOLS)

    all_s1 = s1_new_rows + s1_existing
    all_s2 = s2_new_rows + s2_existing

    write_wecom(WECOM_SHEET1_ID, all_s1)
    write_wecom(WECOM_SHEET2_ID, all_s2)
    print(f'OK:wecom:{date_str}')


def process(source_path, output_path, date_str):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList

    def _make_solid_databar(start_type, start_value, end_type, end_value, color):
        rule = DataBarRule(start_type=start_type, start_value=start_value,
                           end_type=end_type, end_value=end_value,
                           color=color, showValue=True)
        orig = rule.to_tree
        def patched(*args, **kwargs):
            tree = orig(*args, **kwargs)
            db = tree.find('.//dataBar')
            if db is not None:
                db.set('gradient', '0')
            return tree
        rule.to_tree = patched
        return rule

    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    GRAY_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    detail, total = read_source(source_path)
    s1_data, s1_mom_raw, s1_totals_raw = build_sheet1_block(detail, total)
    s2_data = build_sheet2_block(detail, total)

    try:
        wb = load_workbook(output_path)
        ws1 = wb['Sheet1']
        ws2 = wb['Sheet2']
        ws1.insert_rows(1, 14)
        ws2.insert_rows(1, 14)
    except (FileNotFoundError, KeyError):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws2 = wb.create_sheet('Sheet2')

    for ws, headers, d_rows, is_s2 in [
        (ws1, SHEET1_HEADERS, s1_data, False),
        (ws2, SHEET2_HEADERS, s2_data, True)
    ]:
        ws.cell(row=1, column=1, value='本周数据').fill = YELLOW_FILL
        ws.cell(row=1, column=2, value=date_str).fill = YELLOW_FILL
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = GRAY_FONT
            c.fill = GRAY_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center')
        for ri, rd in enumerate(d_rows):
            is_t = (ri == len(d_rows) - 1)
            for ci, v in enumerate(rd, 1):
                c = ws.cell(row=3 + ri, column=ci, value=v)
                c.border = THIN_BORDER
                if is_t:
                    c.fill = GRAY_FILL
                    if ci == 1:
                        c.font = Font(bold=True)
                if is_s2:
                    if ci in {3, 6, 9, 12, 14} and isinstance(v, (int, float)) and v is not None:
                        c.number_format = '0.00%'
                else:
                    if ci in {3, 4, 8, 9, 13, 14} and isinstance(v, (int, float)) and v is not None:
                        c.number_format = '0.00%'

    for ws_enum in [(ws1, SHEET1_HEADERS), (ws2, SHEET2_HEADERS)]:
        ws, hdrs = ws_enum
        for ci, h in enumerate(hdrs, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(str(h)) * 2.5, 14)

    mom_cols = [(4, 'D', 0), (9, 'I', 1), (14, 'N', 2)]
    ws1.conditional_formatting = ws1.conditional_formatting.__class__()

    for ri in range(10):
        for col_idx, col_letter, mom_idx in mom_cols:
            raw_val = s1_mom_raw[ri][mom_idx]
            if raw_val is not None:
                c = ws1.cell(row=3 + ri, column=col_idx)
                c.value = raw_val
                c.number_format = '0.00%'

    for col_idx, col_letter, mom_idx in mom_cols:
        rule = _make_solid_databar('num', -1, 'num', 1, '4472C4')
        ws1.conditional_formatting.add(f'{col_letter}3:{col_letter}12', rule)

    if '图表数据' in wb.sheetnames:
        cd = wb['图表数据']
    else:
        cd = wb.create_sheet('图表数据')
    cd.sheet_state = 'hidden'

    cd.delete_rows(1, cd.max_row)
    cd.cell(1, 1, '周次')
    cd.cell(1, 2, date_str)
    cd.cell(2, 1, '活动')
    cd.cell(2, 2, '本周下发算力')
    cd.cell(2, 3, '本周会话消耗算力')
    cd.cell(2, 4, '本周过期消耗算力')

    for i, act in enumerate(ACTIVITY_ORDER):
        cd.cell(3 + i, 1, act)
        cd.cell(3 + i, 2, float(s1_totals_raw['detail_issued'][i]))
        cd.cell(3 + i, 3, float(s1_totals_raw['detail_consumed'][i]))
        cd.cell(3 + i, 4, float(s1_totals_raw['detail_expired'][i]))

    if '图表' in wb.sheetnames:
        del wb['图表']
    chart_sheet = wb.create_sheet('图表')

    excel_colors = ['4472C4', 'ED7D31', 'A5A5A5']
    chart_titles = ['本周下发算力', '本周会话消耗算力', '本周过期消耗算力']
    n_rows = len(ACTIVITY_ORDER) + 2
    anchors = ['A1', 'J1', 'S1']

    for idx, col_idx in enumerate([2, 3, 4]):
        pie = PieChart()
        pie.title = f'{chart_titles[idx]} ({date_str})'
        pie.height = 10
        pie.width = 14

        data_ref = Reference(cd, min_col=col_idx, min_row=2, max_row=n_rows)
        cat_ref = Reference(cd, min_col=1, min_row=3, max_row=n_rows)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cat_ref)

        series = pie.series[0]
        for pt_idx in range(len(ACTIVITY_ORDER)):
            pt = DataPoint(idx=pt_idx)
            pt.graphicalProperties.solidFill = excel_colors[pt_idx % 3]
            series.data_points.append(pt)

        pie.dataLabels = DataLabelList(showPercent=True, showCatName=False, showVal=False)
        pie.legend.position = 'r'
        chart_sheet.add_chart(pie, anchors[idx])

    wb.save(output_path)

    _inject_negative_color(output_path, 'E15759')

    print(f'OK:{output_path}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--source', required=True)
    parser.add_argument('--output', default=None)
    parser.add_argument('--date', required=True)
    parser.add_argument('--wecom', action='store_true')
    args = parser.parse_args()

    if args.wecom:
        if not args.output:
            print('ERROR: --output (local Excel path) required for --wecom mode as data source')
            sys.exit(1)
        process_wecom(args.source, args.date, args.output)
    else:
        if not args.output:
            print('ERROR: --output required')
            sys.exit(1)
        process(args.source, args.output, args.date)
