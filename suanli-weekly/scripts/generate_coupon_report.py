import sys
import json
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


COLUMNS = [
    '日期',
    '创建知识号券数',
    '运营知识库券数', '运营知识库人数',
    '运营skill券数', '运营skill数',
    '总计人数', '总计券数',
    '知识库激励_发一张的人数', '知识库激励_两张', '知识库激励_三张', '知识库激励_四张',
    'skills激励_发一张的人数', 'skills激励_两张', 'skills激励_三张', 'skills激励_四张',
]

STANDARD_HEADERS = [
    '日期',
    '创建知识号券数',
    '运营知识库券数', '运营知识库人数',
    '运营skill券数', '运营skill人数',
    '总计人数', '总计券数',
]

HEADER_ROW = 1
DATA_START_ROW = 2

SEP1_COL = len(STANDARD_HEADERS) + 1
KNOWLEDGE_LABEL_COL = SEP1_COL + 1
KNOWLEDGE_FIRST_COL = KNOWLEDGE_LABEL_COL + 1
SEP2_COL = KNOWLEDGE_FIRST_COL + 4
SKILLS_LABEL_COL = SEP2_COL + 1
SKILLS_FIRST_COL = SKILLS_LABEL_COL + 1
TOTAL_COLS = SKILLS_FIRST_COL + 3

COL_WIDTHS = {
    1: 12, 2: 14, 3: 12, 4: 16, 5: 14, 6: 16, 7: 12, 8: 12,
    SEP1_COL: 4,
    KNOWLEDGE_LABEL_COL: 14, KNOWLEDGE_FIRST_COL: 10, KNOWLEDGE_FIRST_COL + 1: 10,
    KNOWLEDGE_FIRST_COL + 2: 10, KNOWLEDGE_FIRST_COL + 3: 10,
    SEP2_COL: 4,
    SKILLS_LABEL_COL: 14, SKILLS_FIRST_COL: 10, SKILLS_FIRST_COL + 1: 10,
    SKILLS_FIRST_COL + 2: 10, SKILLS_FIRST_COL + 3: 10,
}

STANDARD_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
LABEL_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

KNOWLEDGE_BUCKETS = ['知识库激励_发一张的人数', '知识库激励_两张', '知识库激励_三张', '知识库激励_四张']
SKILLS_BUCKETS = ['skills激励_发一张的人数', 'skills激励_两张', 'skills激励_三张', 'skills激励_四张']


def is_total_row(date_val):
    if pd.isna(date_val):
        return True
    s = str(date_val).strip()
    return s in ('', '总计', 'nan', 'NaN')


def style_header(cell, fill):
    cell.font = HEADER_FONT
    cell.fill = fill
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def generate(input_path, output_path):
    df = pd.read_excel(input_path)

    df['创建知识号券数'] = (
        df['总计券数'].astype(float)
        - df['运营skill券数'].astype(float)
        - df['运营知识库券数'].astype(float)
    ).astype(int)

    available_cols = [c for c in COLUMNS if c in df.columns]
    df = df[available_cols].copy()

    df['_is_total'] = df['日期'].apply(is_total_row)
    df['_date_sort_key'] = df['日期'].apply(lambda v: 99999999 if is_total_row(v) else (int(float(v)) if not pd.isna(v) else 0))
    df = df.sort_values('_date_sort_key').drop(columns=['_is_total', '_date_sort_key']).reset_index(drop=True)

    daily_df = df[~df['日期'].apply(is_total_row)].reset_index(drop=True)
    n_days = len(daily_df)

    def avg_int(col):
        if n_days == 0 or col not in daily_df.columns:
            return 0
        return int(round(daily_df[col].astype(float).sum() / n_days))

    wb = Workbook()
    ws = wb.active
    ws.title = '算力券周报'

    for i, label in enumerate(STANDARD_HEADERS, 1):
        style_header(ws.cell(row=HEADER_ROW, column=i, value=label), STANDARD_FILL)

    ws.cell(row=HEADER_ROW, column=SEP1_COL, value=None).border = THIN_BORDER

    style_header(ws.cell(row=HEADER_ROW, column=KNOWLEDGE_LABEL_COL, value='知识库激励'), LABEL_FILL)
    for i, label in enumerate(['1张', '2张', '3张', '4张']):
        style_header(ws.cell(row=HEADER_ROW, column=KNOWLEDGE_FIRST_COL + i, value=label), STANDARD_FILL)

    ws.cell(row=HEADER_ROW, column=SEP2_COL, value=None).border = THIN_BORDER

    style_header(ws.cell(row=HEADER_ROW, column=SKILLS_LABEL_COL, value='skills激励'), LABEL_FILL)
    for i, label in enumerate(['1张', '2张', '3张', '4张']):
        style_header(ws.cell(row=HEADER_ROW, column=SKILLS_FIRST_COL + i, value=label), STANDARD_FILL)

    col_mapping = {
        '日期': 1,
        '创建知识号券数': 2,
        '运营知识库券数': 3, '运营知识库人数': 4,
        '运营skill券数': 5, '运营skill数': 6,
        '总计人数': 7, '总计券数': 8,
    }
    display_to_source = {
        '运营skill人数': '运营skill数',
    }

    MAIN_COLS = list(range(1, len(STANDARD_HEADERS) + 1))
    SEP_COLS = [SEP1_COL, SEP2_COL]
    GRADIENT_COLS = (
        {KNOWLEDGE_LABEL_COL, *range(KNOWLEDGE_FIRST_COL, KNOWLEDGE_FIRST_COL + 4)}
        | {SKILLS_LABEL_COL, *range(SKILLS_FIRST_COL, SKILLS_FIRST_COL + 4)}
    )

    data_font = Font(name='Arial', size=11)
    data_align = Alignment(horizontal='center', vertical='center')
    total_font = Font(name='Arial', bold=True, size=11)

    for row_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = DATA_START_ROW + row_idx
        is_total = is_total_row(row['日期'])
        date_val = None if is_total else int(float(row['日期']))

        for col in STANDARD_HEADERS:
            src_col = display_to_source.get(col, col)
            if src_col not in available_cols:
                continue
            col_idx = col_mapping[src_col]
            if col == '日期':
                val = '总计' if is_total else date_val
            else:
                val = row[src_col]
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = total_font if is_total else data_font
            cell.alignment = data_align
            cell.border = THIN_BORDER
            if is_total:
                cell.fill = TOTAL_FILL

        for sep_idx in SEP_COLS:
            sc = ws.cell(row=excel_row, column=sep_idx, value=None)
            sc.border = THIN_BORDER

    grad_start = DATA_START_ROW
    for day_idx, (_, row) in enumerate(daily_df.iterrows()):
        excel_row = grad_start + day_idx
        date_val = int(float(row['日期']))

        kc = ws.cell(row=excel_row, column=KNOWLEDGE_LABEL_COL, value=date_val)
        kc.font = data_font
        kc.alignment = data_align
        kc.border = THIN_BORDER
        for i, col in enumerate(KNOWLEDGE_BUCKETS):
            c = ws.cell(row=excel_row, column=KNOWLEDGE_FIRST_COL + i, value=int(float(row[col])) if col in row else 0)
            c.font = data_font
            c.alignment = data_align
            c.border = THIN_BORDER

        sc = ws.cell(row=excel_row, column=SKILLS_LABEL_COL, value=date_val)
        sc.font = data_font
        sc.alignment = data_align
        sc.border = THIN_BORDER
        for i, col in enumerate(SKILLS_BUCKETS):
            c = ws.cell(row=excel_row, column=SKILLS_FIRST_COL + i, value=int(float(row[col])) if col in row else 0)
            c.font = data_font
            c.alignment = data_align
            c.border = THIN_BORDER

    avg_row = grad_start + n_days
    kavg = ws.cell(row=avg_row, column=KNOWLEDGE_LABEL_COL, value='日均')
    kavg.font = total_font
    kavg.fill = TOTAL_FILL
    kavg.alignment = data_align
    kavg.border = THIN_BORDER
    for i, col in enumerate(KNOWLEDGE_BUCKETS):
        c = ws.cell(row=avg_row, column=KNOWLEDGE_FIRST_COL + i, value=avg_int(col))
        c.font = total_font
        c.fill = TOTAL_FILL
        c.alignment = data_align
        c.border = THIN_BORDER

    savg = ws.cell(row=avg_row, column=SKILLS_LABEL_COL, value='日均')
    savg.font = total_font
    savg.fill = TOTAL_FILL
    savg.alignment = data_align
    savg.border = THIN_BORDER
    for i, col in enumerate(SKILLS_BUCKETS):
        c = ws.cell(row=avg_row, column=SKILLS_FIRST_COL + i, value=avg_int(col))
        c.font = total_font
        c.fill = TOTAL_FILL
        c.alignment = data_align
        c.border = THIN_BORDER

    for sep_idx in SEP_COLS:
        for r in range(DATA_START_ROW, avg_row + 1):
            ws.cell(row=r, column=sep_idx, value=None).border = THIN_BORDER

    for col, w in COL_WIDTHS.items():
        if col <= TOTAL_COLS:
            ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[HEADER_ROW].height = 24
    ws.freeze_panes = f'A{DATA_START_ROW}'
    wb.save(output_path)

    verification = []
    for _, row in daily_df.iterrows():
        part_sum = (
            float(row.get('运营skill券数', 0))
            + float(row.get('运营知识库券数', 0))
            + float(row.get('创建知识号券数', 0))
        )
        total = float(row.get('总计券数', 0))
        verification.append({
            'date': str(int(float(row['日期']))),
            'part_sum': part_sum,
            'total': total,
            'match': abs(part_sum - total) < 0.01,
        })

    avg_check = {
        'knowledge': {c: avg_int(c) for c in KNOWLEDGE_BUCKETS},
        'skills': {c: avg_int(c) for c in SKILLS_BUCKETS},
    }

    result = {
        'output': output_path,
        'rows': len(df),
        'daily_rows': n_days,
        'verification': verification,
        'avg_check': avg_check,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('Usage: python generate_coupon_report.py <input.xlsx> <output.xlsx>', file=sys.stderr)
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
