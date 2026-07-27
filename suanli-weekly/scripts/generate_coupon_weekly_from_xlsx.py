import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
import argparse
import re

WEEKS = [
    ("20260626", "20260702"),
    ("20260703", "20260709"),
    ("20260710", "20260716"),
]

DEFAULT_DATA_DIR = os.path.join(os.getcwd(), "beacon_output")


def parse_weeks(arg):
    weeks = []
    for w in arg.split(","):
        w = w.strip()
        if "-" in w:
            s, e = w.split("-")
            weeks.append((s.strip(), e.strip()))
    return weeks


def parse_sql_people(arg):
    result = {}
    for part in arg.split(","):
        part = part.strip()
        m = re.match(r"(kb|skill|total)\s*=\s*(\d+)", part)
        if m:
            result[m.group(1)] = int(m.group(2))
    return result


def main():
    parser = argparse.ArgumentParser(description="Generate formatted coupon weekly reports from 49d xlsx")
    parser.add_argument("--weeks", type=str, default="",
                        help="Week ranges, comma-separated, e.g. '20260626-20260702,20260703-20260709'")
    parser.add_argument("--input-dir", type=str, default=DEFAULT_DATA_DIR,
                        help="Directory containing the 49d xlsx files (default: <cwd>/beacon_output)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Output directory (default: same as --input-dir)")
    parser.add_argument("--sql-people", type=str, required=True,
                        help="SQL-computed weekly unique people: 'kb=N,skill=M,total=P'. One set per week, or per-week format: '20260626:kb=10,skill=5,total=20|20260703:kb=15,skill=8,total=30'")
    parser.add_argument("--daily-people", type=str, default="",
                        help="SQL-computed daily unique people (replaces G column): '20260605=N,20260608=M,...'")
    args = parser.parse_args()

    data_dir = args.input_dir
    out_dir = args.output_dir or data_dir
    weeks = parse_weeks(args.weeks) if args.weeks else WEEKS

    sql_people_map = {}
    if ":" in args.sql_people:
        for block in args.sql_people.split("|"):
            block = block.strip()
            if ":" in block:
                ws, vals = block.split(":", 1)
                sql_people_map[ws.strip()] = parse_sql_people(vals)
    else:
        parsed = parse_sql_people(args.sql_people)
        for ws, we in weeks:
            sql_people_map[ws] = parsed

    daily_people = {}
    if args.daily_people:
        for part in args.daily_people.split(","):
            part = part.strip()
            if "=" in part:
                d, n = part.split("=", 1)
                daily_people[d.strip()] = int(n.strip())

    global FL_XLSX, KB_XLSX, SK_XLSX, OUT_DIR
    FL_XLSX = os.path.join(data_dir, "charoniwang_copilot算力报表_算力券发放-分发放类型_49d.xlsx")
    KB_XLSX = os.path.join(data_dir, "charoniwang_copilot算力报表_算力券发放-运营知识库激励_49d.xlsx")
    SK_XLSX = os.path.join(data_dir, "charoniwang_copilot算力报表_算力券发放-运营skill激励_49d.xlsx")
    OUT_DIR = out_dir

    for f in [FL_XLSX, KB_XLSX, SK_XLSX]:
        if not os.path.exists(f):
            print(f"ERROR: {f} not found. Run 'coupon_api_to_xlsx.py' first.", file=sys.stderr)
            sys.exit(1)

    outs = []
    for ws, we in weeks:
        sql_people = sql_people_map.get(ws)
        out = gen_week_report(ws, we, sql_people, daily_people)
        outs.append(out)
    print(f"\n=== Generated {len(outs)} weekly reports ===")
    for o in outs:
        print(f"  {o}")


def gen_week_report(week_start, week_end, sql_people=None, daily_people=None):
    START = week_start
    END = week_end
    DATES = [d.strftime("%Y%m%d") for d in pd.date_range(START, END, freq="D")]

    df_fenlei = pd.read_excel(FL_XLSX)
    df_fenlei["imp_date"] = df_fenlei["时间"].dt.strftime("%Y%m%d")
    df_fenlei = df_fenlei[df_fenlei["imp_date"].between(START, END)]

    df_kb = pd.read_excel(KB_XLSX)
    df_kb["imp_date"] = df_kb["日期"].dt.strftime("%Y%m%d")
    df_kb = df_kb[df_kb["imp_date"].between(START, END)]

    df_sk = pd.read_excel(SK_XLSX)
    df_sk["imp_date"] = df_sk["日期"].dt.strftime("%Y%m%d")
    df_sk = df_sk[df_sk["imp_date"].between(START, END)]

    HEADER_BLUE = "4472C4"
    DEEP_BLUE = "2F5496"
    LIGHT_BLUE = "D6E4F0"
    BORDER_COLOR = "B4C6E7"

    wb = Workbook()
    ws = wb.active
    ws.title = "算力券周报"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    deep_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
    total_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)

    headers = [
        "日期", "创建知识号券数", "运营知识库券数", "运营知识库人数",
        "运营skill券数", "运营skill人数", "总计人数", "总计券数",
        "", "知识库激励", "1张", "2张", "3张", "4张",
        "", "skills激励", "1张", "2张", "3张", "4张",
    ]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if h in ("知识库激励", "skills激励"):
            cell.fill = deep_fill
        elif h == "":
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            cell.fill = header_fill
        cell.border = thin_border

    data = {}
    days_in_week = len(DATES)

    for d in DATES:
        row = df_fenlei[df_fenlei["imp_date"] == d]

        def get_metric(reason, metric_col):
            r = row[row["发放类型"] == reason]
            if len(r) == 0:
                return 0
            v = r[metric_col].values[0]
            if pd.isna(v):
                return 0
            return int(v)

        create_num = get_metric("创建知识号", "实际发放算力券（张）\t")
        kb_num = get_metric("运营知识库", "实际发放算力券（张）\t")
        sk_num = get_metric("运营skill", "实际发放算力券（张）\t")
        create_people = get_metric("创建知识号", "实际发放人数")
        kb_people = get_metric("运营知识库", "实际发放人数")
        sk_people = get_metric("运营skill", "实际发放人数")
        total_num = create_num + kb_num + sk_num
        if daily_people and d in daily_people:
            total_people = daily_people[d]
        else:
            total_people = max(create_people, kb_people, sk_people)

        def get_kb_bucket(d, b):
            r = df_kb[(df_kb["imp_date"] == d) & (df_kb["应发算力券梯度（张）"] == b)]
            if len(r) == 0:
                return 0
            v = r["发放人数"].values[0]
            return int(v) if not pd.isna(v) else 0

        def get_sk_bucket(d, b):
            r = df_sk[(df_sk["imp_date"] == d) & (df_sk["应发算力券梯度（张）"] == b)]
            if len(r) == 0:
                return 0
            v = r["发放人数"].values[0]
            return int(v) if not pd.isna(v) else 0

        data[d] = {
            "create_num": create_num, "kb_num": kb_num, "sk_num": sk_num,
            "create_people": create_people, "kb_people": kb_people, "sk_people": sk_people,
            "total_num": total_num, "total_people": total_people,
            "kb_b": [get_kb_bucket(d, b) for b in [1, 2, 3, 4]],
            "sk_b": [get_sk_bucket(d, b) for b in [1, 2, 3, 4]],
        }

    for ri, d in enumerate(DATES, 2):
        r = data[d]
        vals = [
            d, r["create_num"], r["kb_num"], r["kb_people"],
            r["sk_num"], r["sk_people"], r["total_people"], r["total_num"],
            "", d,
            r["kb_b"][0], r["kb_b"][1], r["kb_b"][2], r["kb_b"][3],
            "", d,
            r["sk_b"][0], r["sk_b"][1], r["sk_b"][2], r["sk_b"][3],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    total_row = len(DATES) + 2
    sum_create = sum(data[d]["create_num"] for d in DATES)
    sum_kb_num = sum(data[d]["kb_num"] for d in DATES)
    sum_sk_num = sum(data[d]["sk_num"] for d in DATES)
    sum_total_num = sum_create + sum_kb_num + sum_sk_num

    total_kb_people = sql_people.get("kb", 0)
    total_sk_people = sql_people.get("skill", 0)
    total_total_people = sql_people.get("total", 0)
    people_source = "SQL"

    total_vals = [
        "总计",
        sum_create, sum_kb_num, total_kb_people,
        sum_sk_num, total_sk_people, total_total_people, sum_total_num,
        "", "日均",
    ]

    kb_avgs = [round(sum(data[d]["kb_b"][i] for d in DATES) / days_in_week) for i in range(4)]
    sk_avgs = [round(sum(data[d]["sk_b"][i] for d in DATES) / days_in_week) for i in range(4)]

    total_vals += kb_avgs + ["", "日均"] + sk_avgs

    for ci, v in enumerate(total_vals, 1):
        cell = ws.cell(row=total_row, column=ci, value=v)
        cell.font = bold_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    col_widths = {
        "A": 14, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14,
        "I": 5, "J": 14, "K": 10, "L": 10, "M": 10, "N": 10,
        "O": 5, "P": 14, "Q": 10, "R": 10, "S": 10, "T": 10,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    OUT = f"{OUT_DIR}/ima_算力券周报_{START}_{END}.xlsx"
    wb.save(OUT)

    print(f"\n=== {START}-{END} ===")
    print(f"输出: {OUT}")
    if sql_people:
        print(f"SQL 人数: KB={total_kb_people} SK={total_sk_people} 总计={total_total_people} (来源: {people_source})")
    else:
        print(f"人数来源: {people_source}")
    for d in DATES:
        r = data[d]
        check = r["create_num"] + r["kb_num"] + r["sk_num"]
        match = "OK" if check == r["total_num"] else f"MISMATCH: sum={check} total={r['total_num']}"
        print(f"  {d}: 创建知识号={r['create_num']:>4} KB={r['kb_num']:>4} SK={r['sk_num']:>4} => sum={check} {match}")
        kb_check = sum(r["kb_b"])
        sk_check = sum(r["sk_b"])
        print(f"    KB分桶: {'/'.join(str(x) for x in r['kb_b'])} (sum={kb_check}) | SK分桶: {'/'.join(str(x) for x in r['sk_b'])} (sum={sk_check})")
    print(f"总计: 创建={sum_create} KB券数={sum_kb_num} SK券数={sum_sk_num} 总券数={sum_total_num}")
    print(f"日均 KB: {'/'.join(str(x) for x in kb_avgs)} | SK: {'/'.join(str(x) for x in sk_avgs)}")
    return OUT


if __name__ == "__main__":
    main()
