import sys
import json
from openpyxl import load_workbook

FIELDS = {
    3: '本周下发算力',
    4: '本周会话消耗算力',
    5: '本周过期消耗算力',
}

def process(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active

    buy_row = None
    total_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[1] if len(row) > 1 else None
        if cell and cell.value == '购买算力':
            buy_row = cell.row
        if cell and cell.value == 'TOTAL':
            total_row = cell.row

    if buy_row is None:
        print(json.dumps({'error': '未找到"购买算力"行'}), file=sys.stderr)
        sys.exit(1)
    if total_row is None:
        print(json.dumps({'error': '未找到 TOTAL 行'}), file=sys.stderr)
        sys.exit(1)

    ws.delete_rows(buy_row)
    total_row -= 1

    sum_cols = []
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header and header not in ['sort_no', '活动']:
            sum_cols.append(col_idx)

    for col_idx in sum_cols:
        total_val = 0.0
        for row_idx in range(2, total_row):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                total_val += float(v)
        ws.cell(row=total_row, column=col_idx).value = total_val

    wb.save(output_path)

    activities = []
    for row_idx in range(2, total_row):
        name = ws.cell(row=row_idx, column=2).value
        vals = {}
        for col_idx, field in FIELDS.items():
            v = ws.cell(row=row_idx, column=col_idx).value
            vals[field] = float(v) if v else 0
        activities.append({'name': name, 'values': vals})

    chart_data = {}
    for field_name in FIELDS.values():
        total = sum(a['values'][field_name] for a in activities)
        items = []
        others_val = 0
        for a in activities:
            v = a['values'][field_name]
            pct = v / total * 100 if total > 0 else 0
            if v == 0 or pct < 1:
                others_val += v
            else:
                items.append({'name': a['name'], 'value': v, 'pct': round(pct, 2)})
        if others_val > 0:
            others_pct = others_val / total * 100
            items.append({'name': '其他', 'value': others_val, 'pct': round(others_pct, 2)})
        chart_data[field_name] = {
            'total': total,
            'items': items,
        }

    result = {
        'output': output_path,
        'activities_count': len(activities),
        'charts': chart_data,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('Usage: python process.py <input.xlsx> <output.xlsx>', file=sys.stderr)
        sys.exit(1)
    process(sys.argv[1], sys.argv[2])
