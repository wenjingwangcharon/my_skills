import json
import sys
import os
import asyncio
import re
from pathlib import Path
from playwright.async_api import async_playwright

def find_auth_file():
    paths = [
        os.environ.get("BEACON_AUTH_STATE"),
        Path.cwd() / "beacon_auth_state.json",
        Path.home() / ".workbuddy/skills/beacon-data-fetcher/runtime/beacon_auth_state.json",
        Path.cwd() / "beacon-data-fetcher/runtime/beacon_auth_state.json",
    ]
    for p in paths:
        if p and Path(p).exists():
            return str(Path(p).resolve())
    return None

AUTH_FILE = find_auth_file()
OUTPUT_DIR = Path(os.environ.get("BEACON_OUTPUT_DIR", str(Path.cwd() / "beacon_output")))

if len(sys.argv) < 2:
    print("Usage: python3 beacon_sql_runner.py <sql_file>")
    print("Optional env: BEACON_AUTH_STATE=/path/to/auth.json  BEACON_OUTPUT_DIR=/path/to/output")
    sys.exit(1)

SQL_FILE = Path(sys.argv[1])

def log(s): print(s, flush=True)

async def wait_for_result_table(page, timeout_seconds=120):
    log(f"  Waiting for result table (max {timeout_seconds}s)...")
    for i in range(timeout_seconds // 3):
        await page.wait_for_timeout(3000)
        status = await page.evaluate("""() => {
            const tables = document.querySelectorAll('table');
            let dataRows = 0;
            for (const t of tables) {
                const trs = t.querySelectorAll('tbody tr');
                dataRows += trs.length;
            }
            const bodyText = (document.body?.innerText || '');
            const hasFail = bodyText.includes('查询失败');
            const hasSuccess = bodyText.includes('查询成功');
            const hasComplete = bodyText.includes('运行完成') || bodyText.includes('执行完成');
            return {dataRows, hasFail, hasSuccess, hasComplete};
        }""")
        elapsed = (i + 1) * 3
        log(f"  {elapsed}s: rows={status['dataRows']} fail={status['hasFail']} success={status['hasSuccess']} complete={status['hasComplete']}")
        if status['dataRows'] > 0:
            log("  -> Result table found!")
            return {"result": "success", "rows": status['dataRows']}
        if status['hasFail']:
            log("  -> Query failed!")
            return {"result": "fail"}
    log("  -> Timeout: no result and no failure")
    return {"result": "timeout"}

async def extract_table_data(page):
    data = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        let best = null;
        for (const t of tables) {
            const trs = t.querySelectorAll('tbody tr');
            if (trs.length > 0 && (!best || trs.length > best.rows)) {
                best = {headers: [], rows: [], rowCount: trs.length};
                const ths = t.querySelectorAll('thead th, thead td');
                for (const th of ths) best.headers.push(th.textContent.trim());
                for (const tr of trs) {
                    const row = [];
                    for (const td of tr.querySelectorAll('td, th')) row.push(td.textContent.trim());
                    best.rows.push(row);
                }
            }
        }
        if (!best) {
            const bodyText = document.body?.innerText || '';
            if (bodyText.includes('查询失败')) return {error: 'query_failed', bodyHint: bodyText.substring(0, 2000)};
            return {error: 'no_table'};
        }
        return best;
    }""")
    return data

async def main():
    if not AUTH_FILE:
        log("ERROR: No beacon_auth_state.json found.")
        log("  Set BEACON_AUTH_STATE env var or ensure auth file exists at one of:")
        log("  - <cwd>/beacon_auth_state.json")
        log("  - ~/.workbuddy/skills/beacon-data-fetcher/runtime/beacon_auth_state.json")
        log("  - <cwd>/beacon-data-fetcher/runtime/beacon_auth_state.json")
        log("  Run beacon_qr_server.py to generate auth state.")
        return
    log(f"Auth file: {AUTH_FILE}")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sql_text = SQL_FILE.read_text().strip()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage", "--disable-gpu"]
        )
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            storage_state=AUTH_FILE
        )
        page = await context.new_page()

        log("Step 1: Load page")
        await page.goto(
            "https://beacon.woa.com/datatalk/ima/card?mode=sql",
            wait_until="domcontentloaded", timeout=60000
        )
        for i in range(12):
            await page.wait_for_timeout(5000)
            body = await page.evaluate("() => (document.body?.innerText || '').substring(0, 200)")
            if "SQL模式" in body:
                log(f"  SQL mode ready at {(i+1)*5}s")
                break

        log("Step 2: Close any ant-modal / intro modal / engine-tpl-overlay")
        for _ in range(6):
            count = await page.evaluate("""() => {
                const modals = document.querySelectorAll('.ant-modal-wrap, .dsjtj-el-dialog__wrapper, [class*="intro"][class*="modal"]');
                const overlays = document.querySelectorAll('.engine-tpl-overlay');
                return {modals: modals.length, overlays: overlays.length};
            }""")
            log(f"  Visible: {count}")
            if count['modals'] == 0 and count['overlays'] == 0:
                break
            closed = await page.evaluate("""() => {
                const wraps = document.querySelectorAll('.ant-modal-wrap, .dsjtj-el-dialog__wrapper, [class*="intro"][class*="modal"]');
                for (const w of wraps) {
                    if (w.offsetParent === null) continue;
                    const btns = w.querySelectorAll('button');
                    for (const b of btns) {
                        const t = (b.textContent || '').trim();
                        if (t === '关闭' || t === '知道了' || t === '关 闭' || t === '×') {
                            b.click(); return 'clicked button: ' + t;
                        }
                    }
                    const xBtn = w.querySelector('.ant-modal-close, .dsjtj-el-dialog__close, [class*="close-btn"], [class*="CloseBtn"]');
                    if (xBtn) { xBtn.click(); return 'clicked close icon'; }
                }
                const overlays = document.querySelectorAll('.engine-tpl-overlay');
                for (const ov of overlays) {
                    if (ov.offsetParent === null) continue;
                    const closeBtn = ov.querySelector('[class*="close"], [class*="Close"]');
                    if (closeBtn) { closeBtn.click(); return 'clicked overlay close'; }
                    ov.style.display = 'none';
                    return 'hid overlay';
                }
                return 'no closer found';
            }""")
            log(f"  Close result: {closed}")
            await page.wait_for_timeout(2000)
        await page.screenshot(path=str(OUTPUT_DIR / "00_modal_closed.png"))

        log("Step 3: Select data source")
        await page.click('.category-prefix-trigger', timeout=10000)
        await page.wait_for_timeout(2000)
        clicked = await page.evaluate("""() => {
            const items = document.querySelectorAll('.ant-dropdown-menu-item');
            for (const it of items) {
                const t = (it.textContent || '').trim();
                if (t === 'StarRocks' || t.includes('StarRocks')) {
                    it.click();
                    return 'clicked: ' + t;
                }
            }
            return 'StarRocks not in dropdown';
        }""")
        log(f"  Category select: {clicked}")
        await page.wait_for_timeout(2500)

        await page.click('.source-select', timeout=10000)
        await page.wait_for_timeout(2000)
        await page.keyboard.press("Control+a")
        await page.keyboard.press("Delete")
        await page.wait_for_timeout(500)
        await page.keyboard.type("Ima", delay=50)
        await page.wait_for_timeout(3000)
        await page.screenshot(path=str(OUTPUT_DIR / "01_dropdown_Ima.png"))

        options = await page.evaluate("""() => {
            const items = document.querySelectorAll('.ant-select-item-option, [class*="select-item"]');
            return Array.from(items).filter(el => el.offsetParent !== null).map(el => el.textContent.trim());
        }""")
        log(f"  Visible options: {options}")

        option_result = await page.evaluate("""() => {
            const items = document.querySelectorAll('.ant-select-item-option, [class*="select-item"]');
            for (const el of items) {
                if (el.offsetParent === null) continue;
                const t = (el.textContent || '').trim();
                if (t.includes('Ima') && t.includes('Xingpan')) {
                    el.scrollIntoView({block: 'center'});
                    el.click();
                    return 'clicked: ' + t;
                }
            }
            return 'not found';
        }""")
        log(f"  Option click: {option_result}")
        await page.wait_for_timeout(3000)

        selected_text = await page.evaluate("""() => {
            const sel = document.querySelector('.source-select');
            if (!sel) return 'no .source-select';
            return (sel.textContent || '').trim().substring(0, 200);
        }""")
        log(f"  Selected display: {selected_text}")
        await page.screenshot(path=str(OUTPUT_DIR / "01_datasource_selected.png"))

        log("Step 4: Inject SQL via Monaco paste event")
        await page.click('.monaco-editor .view-lines', timeout=10000, force=True)
        await page.wait_for_timeout(1500)

        clear_result = await page.evaluate("""() => {
            const ta = document.querySelector('.monaco-editor textarea.inputarea');
            if (!ta) return 'no textarea';
            ta.focus();
            ta.setSelectionRange(0, ta.value.length);
            return {ok: true, valLen: ta.value.length, sel: ta.selectionEnd - ta.selectionStart};
        }""")
        log(f"  Pre-clear textarea: {clear_result}")
        await page.keyboard.press("Control+a")
        await page.keyboard.press("Delete")
        await page.wait_for_timeout(500)

        paste_result = await page.evaluate("""(sql) => {
            const ta = document.querySelector('.monaco-editor textarea.inputarea');
            if (!ta) return {ok: false, err: 'no textarea'};
            ta.focus();
            const dt = new DataTransfer();
            dt.setData('text/plain', sql);
            const ev = new ClipboardEvent('paste', {
                clipboardData: dt,
                bubbles: true,
                cancelable: true
            });
            const dispatched = ta.dispatchEvent(ev);
            return {ok: true, dispatched, valAfter: ta.value.length};
        }""", sql_text)
        log(f"  Paste event: {paste_result}")
        await page.wait_for_timeout(3000)

        verify = await page.evaluate("""() => {
            const ta = document.querySelector('.monaco-editor textarea.inputarea');
            const lines = document.querySelectorAll('.monaco-editor .view-line');
            return {
                taLen: ta ? ta.value.length : 0,
                taFirst: ta ? ta.value.substring(0, 100) : '',
                lineCount: lines.length,
                firstLine: lines[0]?.textContent?.trim().substring(0, 80) || '',
                lastLine: lines[lines.length-1]?.textContent?.trim().substring(0, 80) || ''
            };
        }""")
        log(f"  Verify: {verify}")
        await page.screenshot(path=str(OUTPUT_DIR / "02_sql_injected.png"))

        log("Step 5: Click 查询 button")
        await page.evaluate("""() => {
            const all = document.querySelectorAll('button');
            for (const b of all) {
                const t = (b.textContent || '').trim().replace(/\\s+/g, '');
                if (t === '查询' && b.offsetParent !== null && !b.disabled) {
                    b.click();
                    return 'clicked';
                }
            }
            return 'not found';
        }""")
        await page.wait_for_timeout(2000)

        log("Step 6: Wait for query result")
        result = await wait_for_result_table(page, timeout_seconds=120)

        if result["result"] == "success":
            log(f"Step 7: Extracting {result['rows']} data rows")
            await page.screenshot(path=str(OUTPUT_DIR / "03_result_table.png"))
            data = await extract_table_data(page)
            log(f"  Headers: {data.get('headers', [])}")
            log(f"  Row count: {len(data.get('rows', []))}")
            for i, row in enumerate(data.get('rows', [])[:3]):
                log(f"  Row {i}: {row}")

            log("Step 8: Save to Excel")
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "查询结果"
            headers = data.get('headers', [])
            if headers:
                for ci, h in enumerate(headers, 1):
                    ws.cell(row=1, column=ci, value=h)
            for ri, row in enumerate(data.get('rows', []), 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)
            xlsx_path = OUTPUT_DIR / (SQL_FILE.stem + "_result.xlsx")
            wb.save(str(xlsx_path))
            log(f"  Saved: {xlsx_path}")

        elif result["result"] == "fail":
            log("Step 7: Query failed, capturing error")
            await page.screenshot(path=str(OUTPUT_DIR / "03_query_failed.png"))
            err_text = await page.evaluate("""() => (document.body?.innerText || '').substring(0, 5000)""")
            log(f"  Error text: {err_text[:1000]}")

        else:
            log("Step 7: Timeout, saving screenshot")
            await page.screenshot(path=str(OUTPUT_DIR / "03_timeout.png"))
            body_text = await page.evaluate("""() => (document.body?.innerText || '').substring(0, 3000)""")
            log(f"  Body text: {body_text[:1000]}")

        log("DONE")
        await browser.close()

asyncio.run(main())
